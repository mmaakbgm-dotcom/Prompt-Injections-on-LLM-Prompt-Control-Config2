import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

AUDIT_LOG_FILE = "3_2_audit_log.txt"
OUTPUT_FILE = "audit_log_forensics.xlsx"

COLUMNS = [
    "timestamp", "username", "role", "user_input",
    "llm_raw_output", "extracted_sql", "decision", "row_count",
    "clinic_file", "guiding_prompt_hash", "history_hash", "history_len",
    "stage1_model", "stage1_temperature", "stage1_max_tokens",
]

ILLEGAL_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def parse_log_line(line):
    row = {}
    parts = line.split(" | ")
    for part in parts:
        eq_idx = part.find("=")
        if eq_idx == -1:
            continue
        key = part[:eq_idx].strip()
        val = part[eq_idx + 1:].strip()
        row[key] = val
    return row


def main():
    log_file = sys.argv[1] if len(sys.argv) > 1 else AUDIT_LOG_FILE

    try:
        with open(log_file, "r") as f:
            lines = f.readlines()
    except FileNotFoundError:
        print(f"Audit log not found: {log_file}")
        sys.exit(1)

    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        rows.append(parse_log_line(line))

    if not rows:
        print("No log entries found.")
        sys.exit(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "forensics"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = row.get(col_name, "")
            if isinstance(val, str):
                val = ILLEGAL_CHARS_RE.sub("", val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(OUTPUT_FILE)
    print(f"Exported {len(rows)} entries to {OUTPUT_FILE}")
    print(f"Columns: {', '.join(COLUMNS)}")


if __name__ == "__main__":
    main()
